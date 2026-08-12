#!/usr/bin/env python3
"""
Resolve "Unverified Candidate N" placeholders in 2026/counties/*.csv by matching
each candidate column's county-wide vote total against official county totals
(2026/20260519__al__primary__county.csv, cross-checked against the SoS xlsx).

Usage:
    python3 match_candidates.py --dry-run     # report only, no file writes
    python3 match_candidates.py               # apply fixes + write report
    python3 match_candidates.py --county shelby tuscaloosa cullman  # limit scope

See /Users/dwillis/.claude/plans/i-ve-parsed-alabama-2026-spicy-pillow.md for design.
"""

import argparse
import csv
import glob
import os
import re
from collections import Counter, defaultdict
from difflib import SequenceMatcher

REPO = os.path.dirname(os.path.abspath(__file__))
COUNTIES_DIR = os.path.join(REPO, "2026", "counties")
COUNTY_CSV = os.path.join(REPO, "2026", "20260519__al__primary__county.csv")
XLSX_PATH = os.path.join(
    REPO,
    "2026 AL Republican Party Primary Precinct Results",
    "Final Primary Results by County 2026 5.27.26 5.31pm.xlsx",
)
REPORT_PATH = os.path.join(REPO, "2026", "candidate_match_report.md")

SKIP_FILES = {"20260519__al__primary__democratic__precinct.csv"}
NON_CONFORMING = {"20260519__al__primary__franklin__precinct.csv"}  # 4-col, no candidates

PLACEHOLDER_RE = re.compile(r"^Unverified Candidate \d+$")

# --- text normalization -----------------------------------------------------

_TYPO_FIXES = [
    (r"\bgovenor\b", "governor"),
    (r"\bcommissidner\b", "commissioner"),
    (r"\bano industries\b", "and industries"),
    (r"\bcdunty\b", "county"),
    (r"\bsounty\b", "county"),
    (r"\blauoerdale\b", "lauderdale"),
    (r"\bshleby\b", "shelby"),
    (r"\bgop\b", "republican"),
    (r"\bdistric\b", "district"),
    (r"\bcommitttee\b", "committee"),
    (r"\bexec comm\b", "executive committee"),
    (r"\bpl no\b", "place no"),
    (r"\bpl\.?\s*no\b", "place no"),
    (r"\bplace n\b", "place no"),
    (r"\bdist no\b", "district no"),
    (r"\bdist\.?\b", "district"),
    (r"\bunited states senator\b", "u.s. senate"),
    (r"\bunited states representative\b", "u.s. house"),
    (r"\bsenator\b", "u.s. senate"),
    (r"\bstate senator\b", "state senate"),
]


def normalize_office(s):
    s = (s or "").lower().strip()
    s = s.replace(",", " ").replace(".", " ").replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    for pat, repl in _TYPO_FIXES:
        s = re.sub(pat, repl, s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_district(d):
    if d is None:
        return ""
    d = str(d).strip()
    if not d:
        return ""
    # "4.0" -> "4"
    try:
        f = float(d)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return d


def office_similarity(a, b):
    a, b = normalize_office(a), normalize_office(b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a.startswith(b) or b.startswith(a):
        shorter, longer = sorted((len(a), len(b)))
        return 0.85 + 0.15 * (shorter / longer)
    ratio = SequenceMatcher(None, a, b).ratio()
    ta, tb = set(a.split()), set(b.split())
    if ta and tb:
        token_ratio = len(ta & tb) / len(ta | tb)
        ratio = max(ratio, token_ratio)
    return ratio


def name_similarity(a, b):
    a = re.sub(r"[^a-z ]", "", (a or "").lower())
    b = re.sub(r"[^a-z ]", "", (b or "").lower())
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


# --- executive-committee "Place N" family grouping --------------------------

_PLACE_SUFFIX_RE = re.compile(
    r"\s*,?\s*(place\s*(no\.?)?\s*\d+|district\s*(no\.?)?\s*\d+)\s*$"
)


def family_key(office_norm):
    """Strip a trailing 'Place N' / 'District No. N' suffix, used to pool
    executive-committee contests that were split across canvass columns."""
    return _PLACE_SUFFIX_RE.sub("", office_norm).strip()


def is_committee_office(office_norm):
    return "executive committee" in office_norm


# --- authority table ---------------------------------------------------------


class AuthorityTable:
    """county -> list of contests: {office_raw, office_norm, district, party, candidates: [(name, total), ...]}"""

    def __init__(self):
        self.by_county = defaultdict(list)
        self._load_county_csv()
        self._load_xlsx()
        self._build_pooled_committee_contests()

    def _add_contest(self, county, office_raw, district, party, candidates, source):
        if not candidates:
            return
        self.by_county[county].append(
            {
                "office_raw": office_raw,
                "office_norm": normalize_office(office_raw),
                "district": normalize_district(district),
                "party": party,
                "candidates": candidates,  # list of (name, total) in file order
                "source": source,
                "pooled": False,
            }
        )

    def _load_county_csv(self):
        if not os.path.exists(COUNTY_CSV):
            return
        rows = list(csv.DictReader(open(COUNTY_CSV, newline="", encoding="utf-8")))
        grouped = defaultdict(list)
        order = []
        for r in rows:
            key = (r["county"], r["office"], r.get("district", ""), r["party"])
            if key not in grouped:
                order.append(key)
            grouped[key].append((r["candidate"], _to_int(r["votes"])))
        for key in order:
            county, office, district, party = key
            self._add_contest(county, office, district, party, grouped[key], "county.csv")

    def _load_xlsx(self):
        if not os.path.exists(XLSX_PATH):
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            return
        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name == "Summary":
                continue
            county = sheet_name.strip()
            ws = wb[sheet_name]
            contest_office = None
            contest_candidates = []

            def flush():
                if contest_office and contest_candidates:
                    district = _extract_district(contest_office)
                    if not self._has_similar(county, contest_office, district, "REP"):
                        self._add_contest(
                            county, contest_office, district, "REP", list(contest_candidates), "xlsx"
                        )

            for row in ws.iter_rows(values_only=True):
                vals = [c for c in row if c is not None]
                if not vals:
                    continue
                first = str(vals[0]).strip()
                if len(vals) >= 2 and str(vals[1]).strip() == "Votes":
                    flush()
                    contest_office = first
                    contest_candidates = []
                elif first.lower() == "total":
                    continue
                elif contest_office is not None and len(vals) >= 2:
                    votes = _to_int(vals[1])
                    if votes is not None:
                        contest_candidates.append((first, votes))
            flush()

    def _has_similar(self, county, office, district, party):
        d = normalize_district(district)
        for c in self.by_county.get(county, []):
            if c["party"] != party:
                continue
            # only treat as a genuinely different contest when BOTH sides carry
            # a non-blank district and they disagree (e.g. two U.S. House seats);
            # a blank on either side (common for xlsx titles that aren't really
            # districted, like committee "Place" names matching \bdistrict\b) is
            # treated as compatible so we don't load a duplicate.
            if c["district"] and d and c["district"] != d:
                continue
            if office_similarity(office, c["office_raw"]) >= 0.75:
                return True
        return False

    def _build_pooled_committee_contests(self):
        for county, contests in list(self.by_county.items()):
            families = defaultdict(list)
            for c in contests:
                if not is_committee_office(c["office_norm"]):
                    continue
                fam = (family_key(c["office_norm"]), c["party"])
                families[fam].append(c)
            for (fam_office, party), members in families.items():
                if len(members) < 2:
                    continue
                pooled_candidates = []
                for m in members:
                    pooled_candidates.extend(m["candidates"])
                self.by_county[county].append(
                    {
                        "office_raw": fam_office,
                        "office_norm": fam_office,
                        "district": "",
                        "party": party,
                        "candidates": pooled_candidates,
                        "source": "pooled(" + "+".join(m["source"] for m in members) + ")",
                        "pooled": True,
                    }
                )

    def contests_for(self, county, party):
        return [c for c in self.by_county.get(county, []) if c["party"] == party]


def _to_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "")
    if not s or not re.match(r"^-?\d+$", s):
        return None
    return int(s)


def _extract_district(office):
    m = re.search(r"district(?:\s*no\.?)?\s*(\d+)", office, re.I)
    if m:
        return m.group(1)
    m = re.search(r"\b(\d+)(?:st|nd|rd|th)\s+congressional", office, re.I)
    if m:
        return m.group(1)
    return ""


# --- matching a county CSV's contests against the authority table ----------


def build_our_contests(rows):
    """rows: list of dict from a county precinct CSV. Returns list of contest groups:
    {office_raw, district, party, rows: [row,...], totals_by_candidate: {cand: total}}
    keyed by (office_raw, district, party) in first-seen order."""
    grouped = {}
    order = []
    for r in rows:
        key = (r["office"], normalize_district(r["district"]), r["party"])
        if key not in grouped:
            grouped[key] = {
                "office_raw": r["office"],
                "district": key[1],
                "party": r["party"],
                "rows": [],
                "totals_by_candidate": Counter(),
            }
            order.append(key)
        g = grouped[key]
        g["rows"].append(r)
        g["totals_by_candidate"][r["candidate"]] += int(r["votes"] or 0)
    return [grouped[k] for k in order]


def total_overlap_score(our_totals, auth_totals):
    c1, c2 = Counter(our_totals), Counter(auth_totals)
    common = sum((c1 & c2).values())
    denom = max(len(our_totals), 1)
    return common / denom


def find_best_authority_contest(our_contest, auth_contests):
    party = our_contest["party"]
    candidates = [c for c in auth_contests if c["party"] == party]
    if not candidates:
        return None, 0.0

    our_totals = list(our_contest["totals_by_candidate"].values())
    our_district = our_contest["district"]

    best, best_score = None, -1.0
    for c in candidates:
        name_score = office_similarity(our_contest["office_raw"], c["office_raw"])
        auth_totals = [t for _, t in c["candidates"]]
        overlap_score = total_overlap_score(our_totals, auth_totals)

        district_penalty = 0.0
        if our_district and c["district"] and our_district != c["district"]:
            district_penalty = 0.5
        elif our_district and not c["district"] and not c.get("pooled"):
            # our contest has a district but candidate contest doesn't advertise one;
            # only a mild penalty since many offices don't carry district in the auth table
            district_penalty = 0.05

        if name_score >= 0.95:
            # near-exact office name match: trust it even if OCR garbled every
            # vote total in the contest (overlap_score would otherwise be 0)
            combined = 0.75 + 0.25 * overlap_score - district_penalty
        else:
            combined = 0.4 * name_score + 0.6 * overlap_score - district_penalty
        if combined > best_score:
            best_score, best = combined, c

    return best, best_score


MATCH_THRESHOLD = 0.45


def resolve_contest(our_contest, auth_contest, county, report):
    """Mutates our_contest['rows'] candidate names in place (via a returned rename map).
    Returns dict of {old_row_index_in_group: new_candidate_name} and appends report lines."""
    auth_candidates = list(auth_contest["candidates"])  # [(name, total)]
    claimed = set()  # indices into auth_candidates already assigned to an our-row

    # our rows in this contest, in original order, with running per-row identity
    # (a row is (candidate_label, vote_count) but multiple precinct rows share the
    # same candidate_label; we operate at the *column* level: one entry per distinct
    # candidate_label with its county-wide total)
    our_columns = list(our_contest["totals_by_candidate"].items())  # [(label, total)]

    renames = {}  # old_label -> new_label

    def auth_total_index_by_exact(total, exclude=()):
        matches = [
            i
            for i, (n, t) in enumerate(auth_candidates)
            if t == total and i not in claimed and i not in exclude
        ]
        return matches

    # Pass B: re-verify already-named candidates
    still_unresolved = []
    for label, total in our_columns:
        if PLACEHOLDER_RE.match(label):
            still_unresolved.append((label, total))
            continue
        # find the auth candidate this name most resembles
        best_i, best_ns = None, 0.0
        for i, (n, t) in enumerate(auth_candidates):
            ns = name_similarity(label, n)
            if ns > best_ns:
                best_ns, best_i = ns, i
        if best_i is not None and best_ns >= 0.6:
            auth_name, auth_total = auth_candidates[best_i]
            if auth_total == total:
                claimed.add(best_i)
                report.append(
                    f"  - OK: `{label}` verified ({total} votes matches official {auth_name})"
                )
                continue
            # mismatch: is there an exact-total match elsewhere (a swap candidate)?
            swap_candidates = auth_total_index_by_exact(total, exclude={best_i})
            if len(swap_candidates) == 1:
                j = swap_candidates[0]
                new_name, _ = auth_candidates[j]
                if new_name != label:
                    renames[label] = new_name
                    claimed.add(j)
                    report.append(
                        f"  - AUTO-FIX: `{label}` ({total} votes) renamed to `{new_name}` "
                        f"(exact total match; official {label}~{auth_name} total was {auth_total})"
                    )
                    continue
            report.append(
                f"  - MISMATCH: `{label}` has {total} votes but official {auth_name} shows {auth_total} "
                f"(left unchanged)"
            )
            claimed.add(best_i)
            continue
        still_unresolved.append((label, total))

    # Pass A: exact unique match for placeholders (and any unnamed-but-unresolved column)
    remaining_after_a = []
    for label, total in still_unresolved:
        exact = auth_total_index_by_exact(total)
        if len(exact) == 1:
            i = exact[0]
            new_name, _ = auth_candidates[i]
            renames[label] = new_name
            claimed.add(i)
            report.append(f"  - MATCHED: `{label}` ({total} votes) -> `{new_name}` (exact total)")
        else:
            remaining_after_a.append((label, total))

    # Pass C: ordered inference for what's left, if counts line up
    unclaimed_auth = [
        (i, n, t) for i, (n, t) in enumerate(auth_candidates) if i not in claimed
    ]
    if remaining_after_a and len(remaining_after_a) == len(unclaimed_auth):
        ok = True
        pairs = list(zip(remaining_after_a, unclaimed_auth))
        for (label, total), (i, n, t) in pairs:
            tol = max(5, round(0.02 * max(total, t)))
            if abs(total - t) > tol:
                ok = False
                break
        if ok:
            for (label, total), (i, n, t) in pairs:
                renames[label] = n
                claimed.add(i)
                report.append(
                    f"  - INFERRED: `{label}` ({total} votes) -> `{n}` (official {t}, ordered match)"
                )
            remaining_after_a = []

    for label, total in remaining_after_a:
        report.append(
            f"  - UNRESOLVED: `{label}` ({total} votes) — no official match found "
            f"(possible artifact/write-in; left as placeholder)"
        )

    unclaimed_names = [n for i, n, t in unclaimed_auth if i not in claimed]
    if unclaimed_names:
        report.append(f"  - (unclaimed official candidates: {', '.join(unclaimed_names)})")

    return renames


# --- ballot authority (order-based, no vote totals) -------------------------
#
# The sample ballots give each office's candidates in printed (ballot) order
# but NO vote totals, so they can't be joined on totals the way the county CSV
# and xlsx are. They matter because they cover strictly-local races (county
# party committee, sheriff, board of education) that never appear in the
# county CSV at all — exactly the contests resolve_contest() leaves as
# "Unverified Candidate N" placeholders. Since the canvass columns are in the
# same ballot order the CSV columns preserve, an all-placeholder contest whose
# column count equals the ballot's candidate count can be named positionally.
#
# This is NOT self-verifying like the totals join (there is no number to check
# against), so it is applied conservatively — only when column count == ballot
# candidate count, and any already-named columns must stay consistent with the
# ballot at their positions or the whole alignment is rejected — and every
# assignment is logged as BALLOT so it is auditable.

_FOR_PREFIX_RE = re.compile(r"^\s*for\s+", re.I)


def _strip_for(office):
    return _FOR_PREFIX_RE.sub("", office or "").strip()


def _office_numbers(*parts):
    nums = []
    for p in parts:
        nums += [int(n) for n in re.findall(r"\d+", str(p or ""))]
    return sorted(nums)


class BallotAuthority:
    """Lazily loads, per county, the {party: {office_raw: [candidate,...]}}
    structure extracted from that county's sample ballots (ballot_extract.py).

    Disabled by default: extraction makes network + vision-API calls. When
    enabled but extraction fails (no ballot, llm not configured, etc.) the
    county simply yields no ballot data and the note is surfaced in the report
    — never an error that aborts the run.
    """

    def __init__(self, enabled=False, model=None):
        self.enabled = enabled
        self.model = model
        self._cache = {}   # county_name -> {party: {office_raw: [cands]}}
        self._notes = {}   # county_name -> str (failure reason, if any)

    def for_county(self, county_name):
        if not self.enabled:
            return {}
        if county_name in self._cache:
            return self._cache[county_name]
        try:
            import ballot_extract
            kwargs = {"model": self.model} if self.model else {}
            data = ballot_extract.extract_county(county_name, **kwargs)
        except Exception as e:  # noqa: BLE001 - any failure -> skip this county's ballots
            self._cache[county_name] = {}
            self._notes[county_name] = f"{type(e).__name__}: {e}"
            return {}
        self._cache[county_name] = data
        return data

    def note(self, county_name):
        return self._notes.get(county_name)


def find_ballot_candidates(oc, ballot_party_offices):
    """Pick the ballot office whose candidate list should name this contest.

    ballot_party_offices: {office_raw: [candidate, ...]} for the contest's party.
    Ranks by office-name similarity (after dropping the ballot's "FOR " prefix)
    and requires the CSV office's embedded numbers to be a subset of the ballot
    office's numbers — so a truncated CSV title with no place number still
    matches its single ballot office, while District No. 1 can't grab District
    No. 3's candidates. Returns (office_raw, [candidates], score) or None.
    """
    csv_nums = set(_office_numbers(oc["office_raw"], oc["district"]))
    survivors = []
    for boffice, cands in ballot_party_offices.items():
        bnums = set(_office_numbers(boffice))
        if not csv_nums.issubset(bnums):
            continue
        score = office_similarity(oc["office_raw"], _strip_for(boffice))
        if score < 0.55:
            continue
        survivors.append((boffice, cands, bnums, score))
    if not survivors:
        return None
    # Rank by name similarity, with a SMALL bonus for an exact number-set match.
    # The bonus (0.03) only breaks near-ties — e.g. "District No. 3" vs the
    # ballot's District 1 and District 3, which score identically on name (they
    # differ by one digit) so the exact {3}=={3} match decides. It must stay
    # smaller than any real name gap so it can't override similarity: "GOP
    # Executive Committee, District 1" (0.86) must still beat "County Commission,
    # District 1" (0.74) even though the latter's {1} exactly matches. Applying
    # the 0.55 similarity floor first is also essential: a truncated CSV office
    # with no numbers ("...Lawrence Count") would otherwise have its empty
    # number set "exactly match" every zero-number ballot office and win.
    EXACT_BONUS = 0.03
    boffice, cands, _, score = max(
        survivors, key=lambda s: s[3] + (EXACT_BONUS if csv_nums == s[2] else 0.0)
    )
    return boffice, cands, score


BALLOT_NAME_CONSISTENCY = 0.6


def resolve_from_ballot(oc, renames, ballot_party_offices, report):
    """Assign remaining placeholder columns from the ballot, positionally.

    `renames` is the totals-based rename map already computed for this contest;
    this function adds ballot-based entries to it in place and returns the count
    of new ballot renames. Applied only when the contest's full column count
    equals the matched ballot office's candidate count, and any already-named
    column stays consistent with the ballot name at its position.
    """
    # current per-column label, in column (== ballot) order, after totals renames
    columns = [renames.get(label, label) for label in oc["totals_by_candidate"]]
    placeholder_positions = [i for i, label in enumerate(oc["totals_by_candidate"])
                             if PLACEHOLDER_RE.match(renames.get(label, label))]
    if not placeholder_positions:
        return 0

    match = find_ballot_candidates(oc, ballot_party_offices)
    if match is None:
        report.append("  - BALLOT: no matching ballot office for the remaining placeholder(s)")
        return 0
    boffice, cands, score = match

    if len(columns) != len(cands):
        report.append(
            f"  - BALLOT: \"{boffice}\" has {len(cands)} candidate(s) but contest has "
            f"{len(columns)} column(s) — counts differ, left as placeholder(s)"
        )
        return 0

    # guard: every already-named column must be consistent with the ballot name
    # at its position, or the positional alignment is untrustworthy.
    for i, label in enumerate(columns):
        if PLACEHOLDER_RE.match(label):
            continue
        if name_similarity(label, cands[i]) < BALLOT_NAME_CONSISTENCY:
            report.append(
                f"  - BALLOT: alignment rejected — named column `{label}` at position {i+1} "
                f"disagrees with ballot `{cands[i]}` (matched \"{boffice}\", score {score:.2f})"
            )
            return 0

    n = 0
    labels = list(oc["totals_by_candidate"])
    for i in placeholder_positions:
        orig_label = labels[i]
        new_name = cands[i]
        renames[orig_label] = new_name
        report.append(
            f"  - BALLOT: `{orig_label}` -> `{new_name}` (position {i+1} of "
            f"\"{boffice}\", score {score:.2f})"
        )
        n += 1
    return n


def process_county_file(path, authority, report, dry_run, ballots=None):
    county_slug = os.path.basename(path).split("__")[3]
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        return 0

    county_name = None
    # match the county-CSV county slug to the authority table's county names
    candidates = list(authority.by_county.keys())
    best = max(candidates, key=lambda c: office_similarity(county_slug, c), default=None)
    if best and office_similarity(county_slug, best) > 0.5:
        county_name = best
    else:
        report.append(f"## {county_slug}\n  - SKIPPED: could not map to an authority county name\n")
        return 0

    our_contests = build_our_contests(rows)
    total_renames = 0
    county_report = [f"## {county_name} ({os.path.basename(path)})"]
    any_activity = False

    ballot_data = ballots.for_county(county_name) if ballots else {}
    ballot_note = ballots.note(county_name) if ballots else None
    if ballot_note:
        county_report.append(f"- (ballot data unavailable: {ballot_note})")
        any_activity = True

    for oc in our_contests:
        has_placeholder = any(PLACEHOLDER_RE.match(l) for l in oc["totals_by_candidate"])
        auth_contests = authority.contests_for(county_name, oc["party"])
        best_auth, score = find_best_authority_contest(oc, auth_contests)

        contest_lines = []
        if best_auth is not None and score >= MATCH_THRESHOLD:
            renames = resolve_contest(oc, best_auth, county_name, contest_lines)
            header = (f"- {oc['office_raw']} (district {oc['district'] or '-'}, {oc['party']}) "
                      f"matched to \"{best_auth['office_raw']}\" [{best_auth['source']}, score {score:.2f}]")
        else:
            renames = {}
            header = (f"- {oc['office_raw']} (district {oc['district'] or '-'}, {oc['party']}): "
                      f"NO CONTEST MATCH (best score {score:.2f})")

        # Ballot pass: name any placeholder columns the totals-based passes
        # couldn't (chiefly the local races absent from the county CSV).
        if ballot_data and oc["party"] in ballot_data:
            resolve_from_ballot(oc, renames, ballot_data[oc["party"]], contest_lines)

        if renames or contest_lines or (has_placeholder and best_auth is None):
            any_activity = True
            county_report.append(header)
            county_report.extend(contest_lines)

        if renames:
            for r in oc["rows"]:
                if r["candidate"] in renames:
                    r["candidate"] = renames[r["candidate"]]
            total_renames += len(renames)

    if any_activity:
        report.extend(county_report)
        report.append("")

    if total_renames and not dry_run:
        fieldnames = ["county", "precinct", "office", "district", "party", "candidate", "votes"]
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)

    return total_renames


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--county", nargs="*", default=None, help="limit to these county slugs")
    ap.add_argument("--ballots", action="store_true",
                    help="also resolve placeholders from sample ballots (ballot_extract.py). "
                         "Makes network + vision-API calls per county; results are cached. "
                         "Off by default.")
    ap.add_argument("--ballot-model", default=None,
                    help="override the vision model used for ballot extraction "
                         "(default: ballot_extract's own default, claude-haiku-4.5)")
    args = ap.parse_args()

    authority = AuthorityTable()
    ballots = BallotAuthority(enabled=args.ballots, model=args.ballot_model)

    paths = sorted(glob.glob(os.path.join(COUNTIES_DIR, "*.csv")))
    report = ["# Candidate Match Report", ""]
    if args.dry_run:
        report.append("_(dry run — no files were modified)_\n")

    total = 0
    for path in paths:
        fname = os.path.basename(path)
        if fname in SKIP_FILES or fname in NON_CONFORMING:
            report.append(f"## {fname}\n  - SKIPPED (non-conforming / not a county file)\n")
            continue
        slug = fname.split("__")[3]
        if args.county and slug not in args.county:
            continue
        total += process_county_file(path, authority, report, args.dry_run, ballots=ballots)

    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(report))

    print(f"{'Would rename' if args.dry_run else 'Renamed'} {total} candidate columns.")
    print(f"Report written to {REPORT_PATH}")


if __name__ == "__main__":
    main()
