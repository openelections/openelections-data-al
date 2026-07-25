#!/usr/bin/env python3
"""
Convert the 2026 Alabama primary XLSX workbooks into a single OpenElections-format
county-level CSV.

Output: 2026/20260519__al__primary__county.csv
Schema:  county,office,district,party,candidate,votes

Inputs:
  * "GOP results.xlsx"                       -> party = REP
  * "2026 Democratic Primary Election Results.xlsx" -> party = DEM

The two workbooks use very different layouts, so each has its own parser; the
results are concatenated, sorted, and written to one CSV.

GOP workbook: a "Summary" sheet (statewide totals) plus one sheet per county.
  Each sheet is a sequence of office blocks:
      <office header>, "Votes", "Percentage"
      <candidate>, <votes>, <percentage>
      ...
      "Total", <total votes>, 1.0
      <blank row>
  The "Total" votes-cast row is dropped, and the Summary sheet's statewide
  rollup rows are dropped as well — the output is per-county only (no
  county="Total" rows).

Democratic workbook: a pivoted layout (one row per candidate, one column per
  county plus a TOTAL column). Two sheets:
  * "Candidates" -> Office, District/Jurisdiction, Place, Ballot Name, <counties>, TOTAL
  * "SDEC"       -> Office, District/Place,         Ballot Name, <counties>, TOTAL
  Per-county cells that are blank (None) are omitted (zero-vote rows are not
  emitted, matching the existing county CSVs). The statewide TOTAL column is
  dropped (no county="Total" rows).

Office names are normalized to the OpenElections canonical names where possible;
U.S. House, State Senate and State House get their district number in the
`district` column. County-local / party-committee offices are kept descriptive
(with place/jurisdiction qualifiers appended) and a blank district, per the
"all offices" scope.
"""

import os
import re

import openpyxl
import pandas as pd

OUTPUT = "2026/20260519__al__primary__county.csv"

GOP_INPUT = "GOP results.xlsx"
DEM_INPUT = "2026 Democratic Primary Election Results.xlsx"

# ---------------------------------------------------------------------------
# GOP workbook parsing
# ---------------------------------------------------------------------------

# Exact canonical office-name mappings (also fixes SoS typos like "Govenor").
GOP_OFFICE_MAP = {
    "Govenor": "Governor",
    "Governor": "Governor",
    "Lt. Govenor": "Lieutenant Governor",
    "Senator": "U.S. Senate",
    "Attorney General": "Attorney General",
    "Secretary of State": "Secretary of State",
    "Treasurer": "State Treasurer",
    "Auditor": "State Auditor",
    "Commissioner of Agriculture & Industries": "Commissioner of Agriculture and Industries",
}

# U.S. House: "... Nth Congressional District" (the "Congr" prefix tolerates
# "Congrssional"-style typos in the source).
US_HOUSE_RE = re.compile(r"(\d+)(?:st|nd|rd|th)?\s+Congr", re.IGNORECASE)
STATE_SENATE_RE = re.compile(r"State\s+Senat(?:or|e)", re.IGNORECASE)
STATE_HOUSE_RE = re.compile(r"State\s+Repre", re.IGNORECASE)
# District clause: "District 4", "District No. 10", "District No 50",
# "District No.22", "Distrct 45" (typo). The "i" in Distri?ct is optional.
DISTRICT_RE = re.compile(r"Distri?ct(?:\s+No\.?)?\s*(\d+)", re.IGNORECASE)


def gop_normalize_office(raw):
    """Map a GOP office-block header to (office, district)."""
    name = raw.strip()

    if name in GOP_OFFICE_MAP:
        return GOP_OFFICE_MAP[name], ""

    m = US_HOUSE_RE.search(name)
    if m:
        return "U.S. House", m.group(1)

    if STATE_SENATE_RE.search(name):
        dm = DISTRICT_RE.search(name)
        return "State Senate", (dm.group(1) if dm else "")

    if STATE_HOUSE_RE.search(name):
        dm = DISTRICT_RE.search(name)
        return "State House", (dm.group(1) if dm else "")

    # State Board of Education: the source spells the same race several ways
    # ("Member, State Board of Education, District 6" / "Member State Board of
    # Education, District No. 6" / "State Board of Education, District 8", with
    # and without a comma). Merge them into one office with the district number
    # in the district column. (County boards say "County Board of Education"
    # and so don't match.)
    if "state board of education" in name.lower():
        dm = DISTRICT_RE.search(name)
        return "State Board of Education", (dm.group(1) if dm else "")

    # County-local office: keep the name as-is, no district.
    return name, ""


def gop_county_from_sheet(sheet_name):
    return "Total" if sheet_name.strip().lower() == "summary" else sheet_name.strip()


def gop_is_header(row):
    """True when row is a GOP office-block header: (text, 'Votes', 'Percentage')."""
    if len(row) < 3:
        return False
    c0, c1, c2 = row[0], row[1], row[2]
    return (
        isinstance(c0, str) and c0.strip() != ""
        and isinstance(c1, str) and c1.strip() == "Votes"
        and isinstance(c2, str) and c2.strip() == "Percentage"
    )


def parse_gop_sheet(ws):
    """Yield (county, office, district, candidate, votes) for one GOP sheet."""
    county = gop_county_from_sheet(ws.title)
    current_office = None

    for row in ws.iter_rows(values_only=True):
        if gop_is_header(row):
            current_office = row[0].strip()
            continue
        if current_office is None:
            continue

        c0 = row[0] if len(row) > 0 else None
        c1 = row[1] if len(row) > 1 else None

        if not isinstance(c0, str) or not c0.strip():
            continue  # blank separator
        if c0.strip() == "Total":
            continue  # office votes-cast row
        if isinstance(c1, (int, float)) and not isinstance(c1, bool):
            office, district = gop_normalize_office(current_office)
            yield (county, office, district, c0.strip(), int(c1))


# ---------------------------------------------------------------------------
# Democratic workbook parsing
# ---------------------------------------------------------------------------

DEM_OFFICE_MAP = {
    "United States Senator": "U.S. Senate",
    "United States Representative": "U.S. House",
    "Governor": "Governor",
    "Lieutenant Governor": "Lieutenant Governor",
    "State Senator": "State Senate",
    "State Representative": "State House",
}
DISTRICTED_OFFICES = {"U.S. House", "State Senate", "State House"}


def dem_normalize_office(office_raw, dj, place):
    """Map a Democratic row's Office + District/Jurisdiction + Place to (office, district)."""
    name = office_raw.strip()
    canonical = DEM_OFFICE_MAP.get(name, name)
    office = canonical
    district = ""

    # Districted offices carry an integer district in District/Jurisdiction.
    if canonical in DISTRICTED_OFFICES and isinstance(dj, (int, float)) and not isinstance(dj, bool):
        district = str(int(dj))
    else:
        # A non-integer jurisdiction (e.g. "Greene County" for District Court
        # Judge, or "1 Female" for SDEC) is appended to the office name.
        if dj is not None and str(dj).strip() != "":
            office = f"{office}, {dj}"

    # A Place qualifier (PSC Place 1, District Court Judge Place 7) keeps
    # distinct races distinct.
    if place is not None and str(place).strip() != "":
        office = f"{office}, Place {place}"

    return office, district


def parse_dem_sheet(ws):
    """Yield (county, office, district, candidate, votes) for one Democratic sheet."""
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    ballot_idx = header.index("Ballot Name")
    total_idx = header.index("TOTAL")
    dj_idx = header.index("District/Jurisdiction") if "District/Jurisdiction" in header else None
    place_idx = header.index("Place") if "Place" in header else None
    county_cols = list(range(ballot_idx + 1, total_idx))

    for r in rows[1:]:
        office_raw = r[0]
        if not isinstance(office_raw, str) or not office_raw.strip():
            continue  # blank separator
        cand = r[ballot_idx]
        if not isinstance(cand, str) or not cand.strip():
            continue
        cand = cand.strip()

        dj = r[dj_idx] if dj_idx is not None else None
        place = r[place_idx] if place_idx is not None else None
        office, district = dem_normalize_office(office_raw, dj, place)

        # Per-county rows: omit blank (None) cells (zero-vote rows are not emitted,
        # matching the existing county CSVs).
        for ci in county_cols:
            v = r[ci]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                yield (header[ci], office, district, cand, int(v))

        # Statewide total row from the reported TOTAL column (verified == sum of
        # the per-county cells, so the output stays internally consistent).
        tv = r[total_idx]
        if isinstance(tv, (int, float)) and not isinstance(tv, bool):
            yield ("Total", office, district, cand, int(tv))


# ---------------------------------------------------------------------------
# Combine + write
# ---------------------------------------------------------------------------

def district_key(d):
    try:
        return int(d)
    except (ValueError, TypeError):
        return -1


def main():
    rows = []

    gop_wb = openpyxl.load_workbook(GOP_INPUT, read_only=True, data_only=True)
    for ws in gop_wb.worksheets:
        for row in parse_gop_sheet(ws):
            rows.append((*row, "REP"))

    dem_wb = openpyxl.load_workbook(DEM_INPUT, read_only=True, data_only=True)
    for ws in dem_wb.worksheets:
        for row in parse_dem_sheet(ws):
            rows.append((*row, "DEM"))

    df = pd.DataFrame(rows, columns=["county", "office", "district", "candidate", "votes", "party"])
    df = df[["county", "office", "district", "party", "candidate", "votes"]]

    # No "Total" (statewide rollup) rows — the output is per-county only.
    df = df[df["county"] != "Total"].copy()

    # Sort by county, then office, district (numeric), party, candidate.
    df["_dkey"] = df["district"].map(district_key)
    df = df.sort_values(["county", "office", "_dkey", "party", "candidate"])
    df = df.drop(columns=["_dkey"]).reset_index(drop=True)

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    df.to_csv(OUTPUT, index=False)
    print(f"Wrote {len(df):,} rows to {OUTPUT}")
    print(df.groupby("party").size().to_string())


if __name__ == "__main__":
    main()